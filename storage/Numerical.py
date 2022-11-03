# run "pip install openpyxl" if not install openpyxl yet
# run "python3 Numerical.py" to get an excel file of Numerical questions

import random
from openpyxl import Workbook

NUMBER_OF_QUESTION_PER_LEVEL = 50

def getData():
	plus_data = []
	minus_data = []

	while len(plus_data) < NUMBER_OF_QUESTION_PER_LEVEL * 8:
		number_1 = random.randint(1001, 9999)
		number_2 = random.randint(1001, 9999)
		if [number_1, number_2] in plus_data or number_1 % 10 == 0 or number_2 % 10 == 0:
			continue
		plus_data.append([number_1, number_2])

	while len(minus_data) < NUMBER_OF_QUESTION_PER_LEVEL * 8:
		number_1 = random.randint(1001, 9999)
		number_2 = random.randint(1001, 9999)
		if [number_1, number_2] in minus_data or number_1 % 10 == 0 or number_2 % 10 == 0 or abs(number_1 - number_2) < 1100:
			continue
		if number_1 > number_2:
			minus_data.append([number_1, number_2])
		else:
			minus_data.append([number_2, number_1])
	return plus_data, minus_data

def generateResult(result, dif_1, dif_2):
	random_result = random.sample([round(result * (1 + dif_1)), round(result * (1 + dif_2))], 2)
	return random_result[0], random_result[1], round(result * (1 + dif_1))

if __name__ == "__main__":
	plus_data, minus_data = getData()
	wb = Workbook()
	ws =  wb.active
	ws.title = "Sheet 1"

	row = 1
	level = 1

	for data in plus_data:
		ws['A' + str(row)] = 'Numerical'
		ws['B' + str(row)] = str(data[0]) + ' + ' + str(data[1])
		result = data[0] + data[1]

		if level == 1:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.02, 0.04)
		elif level == 3:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.02, -0.04)
		elif level == 5:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.01, 0.02)
		elif level == 7:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.01, -0.02)
		elif level == 9:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.005, 0.01)
		elif level == 11:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.005, -0.01)
		elif level == 13:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.01, -0.02)
		elif level == 15:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.01, 0.02)

		ws['F' + str(row)] = level

		if row % NUMBER_OF_QUESTION_PER_LEVEL == 0:
			level = level + 2

		row = row + 1


	level = 2

	for data in minus_data:
		ws['A' + str(row)] = 'Numerical'
		ws['B' + str(row)] = str(data[0]) + ' - ' + str(data[1])
		result = data[0] - data[1]

		if level == 2:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.02, 0.04)
		elif level == 4:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.02, -0.04)
		elif level == 6:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.01, 0.02)
		elif level == 8:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.01, -0.02)
		elif level == 10:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.005, 0.01)
		elif level == 12:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.005, -0.01)
		elif level == 14:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, 0.01, -0.02)
		elif level == 16:
			ws['C' + str(row)], ws['D' + str(row)], ws['E' + str(row)] = generateResult(result, -0.01, 0.02)

		ws['F' + str(row)] = level

		if row % NUMBER_OF_QUESTION_PER_LEVEL == 0:
			level = level + 2

		row = row + 1

	wb.save(filename = 'Numerical.xlsx')
